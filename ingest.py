#!/usr/bin/env python3
"""
Project Anton Egon - Phase 1: RAG Pipeline
Document ingestion script for PDF, DOCX, XLS files with metadata support
"""

import os
import sys
from pathlib import Path
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
import asyncio
import aiofiles

from PyPDF2 import PdfReader
from docx import Document
import openpyxl

from chromadb import Client, Collection
from chromadb.config import Settings
from sentence_transformers import SentenceTransformer
from loguru import logger

# Configuration
VAULT_DIR = Path("vault")
MEMORY_DIR = Path("memory")
CHROMA_PERSIST_DIR = MEMORY_DIR / "chroma_db"
EMBEDDING_MODEL = "sentence-transformers/all-MiniLM-L6-v2"
CHUNK_SIZE = 500
CHUNK_OVERLAP = 50


class DocumentIngestor:
    """Handle document ingestion and embedding creation"""
    
    def __init__(self, persist_directory: Path = CHROMA_PERSIST_DIR):
        """Initialize ChromaDB and embedding model"""
        self.persist_directory = persist_directory
        self.persist_directory.mkdir(parents=True, exist_ok=True)
        
        # Initialize ChromaDB client with persistence
        self.chroma_client = Client(Settings(
            chroma_db_impl="duckdb+parquet",
            persist_directory=str(self.persist_directory)
        ))
        
        # Initialize embedding model
        logger.info(f"Loading embedding model: {EMBEDDING_MODEL}")
        self.embedding_model = SentenceTransformer(EMBEDDING_MODEL)
        
        # Get or create collections for each vault category
        self.collections = {}
        for category in ["internal", "client", "general"]:
            collection_name = f"anton_egon_{category}"
            try:
                self.collections[category] = self.chroma_client.get_or_create_collection(
                    name=collection_name,
                    metadata={"hnsw:space": "cosine"}
                )
                logger.info(f"Collection '{collection_name}' ready")
            except Exception as e:
                logger.error(f"Failed to create collection '{collection_name}': {e}")
                raise
    
    def _extract_text_from_pdf(self, file_path: Path) -> str:
        """Extract text from PDF file"""
        try:
            reader = PdfReader(file_path)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logger.error(f"Failed to extract text from PDF {file_path}: {e}")
            return ""
    
    def _extract_text_from_docx(self, file_path: Path) -> str:
        """Extract text from DOCX file"""
        try:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            logger.error(f"Failed to extract text from DOCX {file_path}: {e}")
            return ""
    
    def _extract_text_from_xlsx(self, file_path: Path) -> str:
        """Extract text from XLSX file"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"--- Sheet: {sheet_name} ---\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    text += row_text + "\n"
            return text
        except Exception as e:
            logger.error(f"Failed to extract text from XLSX {file_path}: {e}")
            return ""
    
    def _chunk_text(self, text: str, chunk_size: int = CHUNK_SIZE, 
                   chunk_overlap: int = CHUNK_OVERLAP) -> List[str]:
        """Split text into overlapping chunks"""
        words = text.split()
        chunks = []
        
        for i in range(0, len(words), chunk_size - chunk_overlap):
            chunk = " ".join(words[i:i + chunk_size])
            if chunk.strip():
                chunks.append(chunk)
        
        return chunks
    
    def _create_metadata(self, file_path: Path, category: str, 
                        chunk_index: int, total_chunks: int) -> Dict[str, Any]:
        """Create metadata for a document chunk"""
        return {
            "source_file": str(file_path.name),
            "source_path": str(file_path),
            "category": category,
            "chunk_index": chunk_index,
            "total_chunks": total_chunks,
            "ingested_at": datetime.now(timezone.utc).isoformat(),
            "file_size": file_path.stat().st_size,
            "file_modified": datetime.fromtimestamp(
                file_path.stat().st_mtime, timezone.utc
            ).isoformat()
        }
    
    def _get_file_category(self, file_path: Path) -> Optional[str]:
        """Determine vault category based on file location"""
        parent = file_path.parent
        if parent.name == "internal":
            return "internal"
        elif parent.name == "client":
            return "client"
        elif parent.name == "general":
            return "general"
        return None
    
    def ingest_file(self, file_path: Path) -> bool:
        """Ingest a single file into the RAG system"""
        logger.info(f"Ingesting file: {file_path}")
        
        # Determine category
        category = self._get_file_category(file_path)
        if not category:
            logger.warning(f"File {file_path} is not in a recognized vault category")
            return False
        
        # Extract text based on file type
        file_ext = file_path.suffix.lower()
        if file_ext == ".pdf":
            text = self._extract_text_from_pdf(file_path)
        elif file_ext == ".docx":
            text = self._extract_text_from_docx(file_path)
        elif file_ext in [".xlsx", ".xls"]:
            text = self._extract_text_from_xlsx(file_path)
        else:
            logger.warning(f"Unsupported file type: {file_ext}")
            return False
        
        if not text.strip():
            logger.warning(f"No text extracted from {file_path}")
            return False
        
        # Chunk the text
        chunks = self._chunk_text(text)
        logger.info(f"Created {len(chunks)} chunks from {file_path}")
        
        # Create embeddings
        embeddings = self.embedding_model.encode(chunks, show_progress_bar=True)
        
        # Prepare metadata and IDs
        ids = []
        metadatas = []
        for i, chunk in enumerate(chunks):
            chunk_id = f"{file_path.stem}_chunk_{i}_{datetime.now().timestamp()}"
            ids.append(chunk_id)
            metadatas.append(self._create_metadata(file_path, category, i, len(chunks)))
        
        # Add to ChromaDB
        collection = self.collections[category]
        collection.add(
            documents=chunks,
            embeddings=embeddings.tolist(),
            metadatas=metadatas,
            ids=ids
        )
        
        # Persist to disk
        self.chroma_client.persist()
        
        logger.info(f"Successfully ingested {file_path} ({len(chunks)} chunks)")
        return True
    
    def ingest_directory(self, directory: Path) -> int:
        """Ingest all supported files in a directory"""
        logger.info(f"Scanning directory: {directory}")
        
        supported_extensions = {".pdf", ".docx", ".xlsx", ".xls"}
        files = [f for f in directory.rglob("*") if f.is_file() and f.suffix.lower() in supported_extensions]
        
        if not files:
            logger.warning(f"No supported files found in {directory}")
            return 0
        
        logger.info(f"Found {len(files)} files to ingest")
        
        success_count = 0
        for file_path in files:
            if self.ingest_file(file_path):
                success_count += 1
        
        logger.info(f"Ingestion complete: {success_count}/{len(files)} files successful")
        return success_count
    
    def purge_old_documents(self, days: int = 30) -> int:
        """Remove documents older than specified days"""
        logger.info(f"Purging documents older than {days} days")
        
        cutoff_date = datetime.now(timezone.utc).timestamp() - (days * 86400)
        total_removed = 0
        
        for category, collection in self.collections.items():
            # Get all documents
            results = collection.get()
            
            if not results or not results["ids"]:
                continue
            
            ids_to_remove = []
            for doc_id, metadata in zip(results["ids"], results["metadatas"]):
                ingested_at = metadata.get("ingested_at", "")
                if ingested_at:
                    try:
                        ingested_timestamp = datetime.fromisoformat(ingested_at.replace("Z", "+00:00")).timestamp()
                        if ingested_timestamp < cutoff_date:
                            ids_to_remove.append(doc_id)
                    except Exception as e:
                        logger.warning(f"Failed to parse date for {doc_id}: {e}")
            
            if ids_to_remove:
                collection.delete(ids=ids_to_remove)
                total_removed += len(ids_to_remove)
                logger.info(f"Removed {len(ids_to_remove)} documents from {category} collection")
        
        if total_removed > 0:
            self.chroma_client.persist()
        
        logger.info(f"Purge complete: {total_removed} documents removed")
        return total_removed


async def main():
    """Main async entry point"""
    logger.add("logs/ingest_{time}.log", rotation="10 MB")
    logger.info("Starting document ingestion pipeline")
    
    # Create ingestor
    ingestor = DocumentIngestor()
    
    # Ingest all vault directories
    for category in ["internal", "client", "general"]:
        vault_path = VAULT_DIR / category
        if vault_path.exists():
            logger.info(f"\n{'='*50}")
            logger.info(f"Processing {category} vault")
            logger.info(f"{'='*50}")
            ingestor.ingest_directory(vault_path)
    
    logger.info("Ingestion pipeline complete")


if __name__ == "__main__":
    asyncio.run(main())
